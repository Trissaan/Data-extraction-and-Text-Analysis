import nltk
import pandas as pd
import numpy as np
import matplotlib.pyplot as plt
import re
from urllib.request import urlopen,Request
from nltk.tokenize import word_tokenize
from bs4 import BeautifulSoup
import openpyxl
sheet1 = openpyxl.load_workbook("/Users/jasper/Downloads/Output Data Structure.xlsx")
op = sheet1.active
pronouns = ["I", "you", "he", "she", "it", "we", "they", "them", "us", "him", "her", "his", "hers", "its", "theirs",
             "our", "your"]
pro_ex = ["US"]
vowels = "AEIOUY"
for f in range(2,op.max_row+1):
 url = op.cell(row=f,column=2).value
 req = Request(url, headers={'User-Agent': 'Chrome/97.0'})
 html = urlopen(req).read()
 soup = BeautifulSoup(html, 'lxml')
 title = soup.find("title").string
 for i in soup.find_all('div', class_='td-post-content'):
    article = i.get_text()
 sheet = openpyxl.load_workbook("/Users/jasper/Downloads/LoughranMcDonald_MasterDictionary_2020.xlsx")
 s = sheet.active
 ex = title + article
 c = f-1
 d = str(c)
 doc = open("/Users/jasper/Documents/" + d, "w")
 doc.write(ex)
 ex1 = re.sub(r'[^\w\s]', '', ex)
 words = word_tokenize(ex1)
 word_count = len(words)
 a = open("/Users/jasper/Downloads/StopWords_Generic.txt", "r")
 b = a.readlines()
 for i in range(len(b)):
     b[i] = b[i].rstrip('\n')
 fil = []
 for i in words:
     if i not in b:
         fil.append(i)
 word_count_fil = len(fil)
 op.cell(row=f, column=12).value = word_count_fil
 pos_count = 0
 neg_count = 0
 for i in range(word_count):
     for j in range(2, s.max_row):
         if s.cell(row=j, column=1).value == words[i].upper():
             if int(s.cell(row=j, column=8).value) > 0:
                 neg_count += 1
             if int(s.cell(row=j, column=9).value) > 0:
                 pos_count += 1
 op.cell(row=f, column=3).value = pos_count
 op.cell(row=f, column=4).value = neg_count
 pol_score = (pos_count - neg_count) / (pos_count + neg_count) + 0.000001
 op.cell(row=f, column=5).value = pol_score
 sub_score = (pos_count + neg_count) / (word_count_fil) + 0.000001
 op.cell(row=f, column=6).value = sub_score
 char_count = len(ex1)-ex1.count(" ")
 sents = nltk.sent_tokenize(ex)
 sent_count = len(sents)
 avg_word_len = char_count/word_count
 op.cell(row=f, column=15).value = avg_word_len
 word_per_sent = word_count/sent_count
 op.cell(row=f, column=7).value = word_per_sent
 op.cell(row=f, column=10).value = word_per_sent
 pronoun_count = 0
 for pro in range(len(pronouns)):
     pronouns[pro] = pronouns[pro].lower()
 for w in range(len(words)):
    if words[w] in pro_ex:
        pronoun_count-=1
    words[w] = words[w].lower()
    if words[w] in pronouns:
        pronoun_count+=1
    words[w] = words[w].upper()
 total_syll_count = 0
 complex_words = 0
 for i in words:
    syll_count = 0
    for j in range(len(i)):
        if i[j] in vowels and i[j-1] not in vowels :
            syll_count +=1
    if i[len(i)-2:] in "ED":
            syll_count-=1
    total_syll_count+= syll_count
    if syll_count > 2:
        complex_words += 1
 op.cell(row=f, column=11).value = complex_words
 complex_precent = (complex_words/word_count)*100
 op.cell(row=f, column=8).value = complex_precent
 syll_per_word = total_syll_count/word_count
 op.cell(row=f, column=13).value = syll_per_word
 fog_index = 0.4 * (word_per_sent + complex_precent)
 op.cell(row=f, column=9).value = fog_index
 op.cell(row=f, column=14).value = pronoun_count
 sheet1.save("/Users/jasper/Downloads/Output Data Structure.xlsx")


























